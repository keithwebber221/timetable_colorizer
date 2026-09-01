
# -*- coding: utf-8 -*-
"""
彩色時間表產生器 (Timetable Colorizer) v6
------------------------------------------------
相較 v5 的新增功能：
  橫向列印格式現在會一併套用：
    - 全部字型改為 Times New Roman
    - 全部字體大小改為 12pt（保留原有的粗體/斜體/顏色等其他樣式）
    - 所有欄寬統一設為 16

執行方式：
    pip install streamlit openpyxl pandas
    streamlit run timetable_colorizer_app.py
"""

import io
import re
from copy import copy
from collections import OrderedDict

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="彩色時間表產生器", layout="wide")

# ----------------------------------------------------------------------------
# 常數設定
# ----------------------------------------------------------------------------
NON_LESSON_KEYWORDS = {
    "morning", "recess", "lunch", "roll", "assembly", "call",
    "早會", "小息", "午膳", "點名", "早自習", "自修",
}

ROOM_RE = re.compile(r"^[A-Za-z]{1,2}\d{2,3}(/[A-Za-z]?\d{2,3})?$|^PG$", re.IGNORECASE)
GRADE_TOKEN_RE = re.compile(r"^(\d)")

# 科目代碼 -> 預設組別。注意：SS / SS2 / SS3 是三個獨立退修組別，不可合併。
DEFAULT_GROUP_MAP = {
    "ENG": "英文組", "ENG_F4": "英文組", "CL": "中文組", "CHIN": "中文組",
    "CHIS": "中史組", "PTH": "普通話組", "HIST": "歷史組",
    "MATH": "數學組", "MACO": "數學組",
    "SS": "退修1組", "SS2": "退修2組", "SS3": "退修3組",
    "CES": "常識組", "SCJ": "科學組", "PHY": "科學組",
    "BIO": "科學組", "JSSP": "科學組", "JSSB": "科學組", "CS": "電腦組",
    "AI": "電腦組", "VA": "視藝組", "HEC": "家政組", "PE": "體育組",
    "BM": "商業組", "ACC": "商業組", "ECON": "商業組", "ARD": "設計與科技組",
    "MOED": "德育及公民教育", "CLP001": "跨學科課程", "CLP002": "跨學科課程",
}

GRADE_CN = {"1": "一", "2": "二", "3": "三", "4": "四", "5": "五", "6": "六"}

DEFAULT_PALETTE = [
    "#FFF2CC", "#D9E1F2", "#E2EFDA", "#FCE4D6", "#D9D2E9", "#FFE599",
    "#C9DAF8", "#D0E0E3", "#F4CCCC", "#D5A6BD", "#B6D7A8", "#A2C4C9",
    "#F9CB9C", "#B4A7D6", "#9FC5E8", "#EAD1DC", "#B7B7B7", "#FFE0B2",
    "#C5CAE9", "#DCEDC8",
]

NONLESSON_LABEL_FILL = "F2F2F2"

# 橫向列印格式的字型設定
PRINT_FONT_NAME = "Times New Roman"
PRINT_FONT_SIZE = 12
PRINT_COLUMN_WIDTH = 16


def is_non_lesson(text: str) -> bool:
    if not text:
        return True
    first_tok = text.strip().split()[0].lower()
    return first_tok in NON_LESSON_KEYWORDS


def extract_grade_label(classes_str: str):
    """回傳 (grade_label, has_anomaly)。grade_label: '1'-'6' / '-' (無級別) / '混合'。"""
    tokens = [c.strip() for c in classes_str.split(",") if c.strip()]
    grades = set()
    for tok in tokens:
        m = GRADE_TOKEN_RE.match(tok)
        if m:
            grades.add(m.group(1))
    if len(grades) == 0:
        return "-", False
    if len(grades) == 1:
        return next(iter(grades)), False
    return "混合", True


def parse_cell(text: str):
    """把 '班別 科目 室別' 字串拆成資訊。若非課堂格則回傳 None。"""
    if text is None:
        return None
    t = str(text).strip()
    if not t or is_non_lesson(t):
        return None
    tokens = t.split()
    if len(tokens) >= 2 and ROOM_RE.match(tokens[-1]):
        subject = tokens[-2]
        room = tokens[-1]
        classes = " ".join(tokens[:-2])
    else:
        subject = tokens[-1]
        room = ""
        classes = " ".join(tokens[:-1])

    anomaly = False
    if not re.search(r"[A-Za-z\u4e00-\u9fff]", subject):
        anomaly = True

    if not classes.strip():
        grade_label, g_anom = "-", False
        anomaly = True
    else:
        grade_label, g_anom = extract_grade_label(classes)
        anomaly = anomaly or g_anom

    return {
        "subject": subject,
        "room": room,
        "classes": classes,
        "grade_label": grade_label,
        "anomaly": anomaly,
    }


def default_group_for(subject: str) -> str:
    key = subject.strip()
    if key in DEFAULT_GROUP_MAP:
        return DEFAULT_GROUP_MAP[key]
    if key.upper() in DEFAULT_GROUP_MAP:
        return DEFAULT_GROUP_MAP[key.upper()]
    base = re.sub(r"\d+$", "", key.upper())
    if base in DEFAULT_GROUP_MAP:
        return DEFAULT_GROUP_MAP[base]
    return f"其他-{key}"


def make_group_name(subject: str, grade_label: str) -> str:
    base = default_group_for(subject)
    if grade_label == "-":
        return f"{base}(其他班別)"
    if grade_label == "混合":
        return f"{base}(混合級別⚠)"
    return f"{base}(中{GRADE_CN.get(grade_label, grade_label)})"


def collect_lesson_keys(wb, sheet_names):
    keys = OrderedDict()
    for name in sheet_names:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                parsed = parse_cell(cell.value)
                if not parsed:
                    continue
                k = (parsed["subject"], parsed["grade_label"])
                if k not in keys:
                    keys[k] = {"classes_examples": set(), "anomaly": False}
                keys[k]["classes_examples"].add(parsed["classes"])
                keys[k]["anomaly"] = keys[k]["anomaly"] or parsed["anomaly"]
    return keys


def build_mapping_df(lesson_keys: "OrderedDict"):
    rows = []
    for (subject, grade_label), info in lesson_keys.items():
        group = make_group_name(subject, grade_label)
        examples = ", ".join(sorted(info["classes_examples"]))[:60]
        rows.append({
            "科目代碼": subject,
            "級別": "全級/其他" if grade_label == "-" else (
                "⚠混合" if grade_label == "混合" else f"中{GRADE_CN.get(grade_label, grade_label)}"),
            "組別": group,
            "班別示例": examples,
            "不填色": bool(info["anomaly"]),
            "⚠備註": "格式異常（多級別混合等），已自動跳過上色" if info["anomaly"] else "",
            "_異常": bool(info["anomaly"]),
        })
    return pd.DataFrame(rows)


def assign_default_colors(groups):
    color_map = {}
    for i, g in enumerate(groups):
        color_map[g] = DEFAULT_PALETTE[i % len(DEFAULT_PALETTE)]
    return color_map


def mapping_to_key_lookup(mapping_df):
    key_to_group = {}
    key_to_skip = {}
    inv_map = {v: k for k, v in GRADE_CN.items()}
    for _, row in mapping_df.iterrows():
        subject = row["科目代碼"]
        level_txt = row["級別"]
        if level_txt == "全級/其他":
            grade_label = "-"
        elif "⚠" in str(level_txt):
            grade_label = "混合"
        else:
            grade_label = inv_map.get(level_txt.replace("中", "").strip(),
                                       level_txt.replace("中", "").strip())
        key = (subject, grade_label)
        key_to_group[key] = row["組別"]
        key_to_skip[key] = bool(row["不填色"])
    return key_to_group, key_to_skip


def apply_landscape_print_layout(ws):
    """
    把工作表設定為適合A4橫向列印的版面：
      - 列印方向、縮放、邊界、凍結標題列（版面設置）
      - 全部字型改 Times New Roman、字體大小改 12pt（保留粗體/斜體/顏色）
      - 所有欄寬統一設為 16
    """
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4,
                                   header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True
    ws.print_options.gridLines = False

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row and max_col:
        ws.print_area = f"A1:{ws.cell(row=max_row, column=max_col).coordinate}"
    if max_row and max_row >= 2:
        ws.print_title_rows = "1:2"

    # 統一字型、字體大小（保留原有粗體/斜體/顏色/底線等樣式）
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            old_font = cell.font
            cell.font = Font(
                name=PRINT_FONT_NAME,
                size=PRINT_FONT_SIZE,
                bold=old_font.bold,
                italic=old_font.italic,
                color=old_font.color,
                underline=old_font.underline,
                strike=old_font.strike,
            )

    # 統一欄寬
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = PRINT_COLUMN_WIDTH


def apply_colors_and_export(wb, sheet_names, full_mapping_df, group_color_map,
                             layout_mode="original", legend_sheet_name="顏色對照表"):
    """在原 workbook 上對所選分頁上色，移除其餘分頁，套用版面格式，並加入顏色對照表分頁。"""
    key_to_group, key_to_skip = mapping_to_key_lookup(full_mapping_df)

    for name in sheet_names:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                text = cell.value
                if text is None or str(text).strip() == "":
                    continue
                if is_non_lesson(text):
                    cell.fill = PatternFill(fill_type="solid",
                                             start_color=NONLESSON_LABEL_FILL,
                                             end_color=NONLESSON_LABEL_FILL)
                    continue
                parsed = parse_cell(text)
                if not parsed:
                    continue
                key = (parsed["subject"], parsed["grade_label"])
                if key_to_skip.get(key, False):
                    continue
                group = key_to_group.get(key)
                color = group_color_map.get(group)
                if color:
                    hex_color = color.lstrip("#").upper()
                    cell.fill = PatternFill(fill_type="solid",
                                             start_color=hex_color, end_color=hex_color)

    for name in list(wb.sheetnames):
        if name not in sheet_names:
            del wb[name]

    if layout_mode == "landscape":
        for name in sheet_names:
            apply_landscape_print_layout(wb[name])

    if legend_sheet_name in wb.sheetnames:
        del wb[legend_sheet_name]
    legend_ws = wb.create_sheet(legend_sheet_name)
    legend_ws.append(["組別", "科目代碼", "級別", "顏色(HEX)", "色塊"])
    for c in legend_ws[1]:
        c.font = Font(bold=True)

    r = 2
    colored_df = full_mapping_df[full_mapping_df["不填色"] == False]
    grouped = colored_df.groupby("組別")
    for group, sub in grouped:
        color = group_color_map.get(group, "").lstrip("#").upper()
        subjects_txt = ", ".join(sorted(set(sub["科目代碼"])))
        levels_txt = ", ".join(sorted(set(sub["級別"])))
        legend_ws.cell(row=r, column=1, value=group)
        legend_ws.cell(row=r, column=2, value=subjects_txt)
        legend_ws.cell(row=r, column=3, value=levels_txt)
        legend_ws.cell(row=r, column=4, value=f"#{color}" if color else "")
        swatch = legend_ws.cell(row=r, column=5, value="")
        if color:
            swatch.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
        r += 1

    for col_letter, width in zip("ABCDE", [26, 26, 12, 12, 8]):
        legend_ws.column_dimensions[col_letter].width = width

    return wb


# ----------------------------------------------------------------------------
# Streamlit 介面
# ----------------------------------------------------------------------------
st.title("🎨 彩色時間表產生器 v6")
st.caption("上載時間表 Excel → 選擇老師 → 依「科目＋級別」分組配色 → 選擇版面格式 → 匯出彩色 Excel")

uploaded = st.file_uploader("上載時間表 Excel (.xlsx)", type=["xlsx"])

if uploaded:
    wb = load_workbook(uploaded, data_only=False)
    all_sheets = wb.sheetnames

    st.subheader("① 選擇要匯出的老師（分頁）")
    sheet_names = st.multiselect("老師分頁", options=all_sheets, default=all_sheets[:1])

    if sheet_names:
        lesson_keys = collect_lesson_keys(wb, sheet_names)

        if not lesson_keys:
            st.warning("在所選分頁中找不到可辨識的課堂格式，請檢查時間表格式。")
        else:
            if ("full_mapping_df" not in st.session_state
                    or st.session_state.get("_keys_key") != tuple(lesson_keys.keys())):
                st.session_state["full_mapping_df"] = build_mapping_df(lesson_keys)
                st.session_state["_keys_key"] = tuple(lesson_keys.keys())

            full_df = st.session_state["full_mapping_df"]
            normal_df = full_df[~full_df["_異常"]].drop(columns=["_異常"]).reset_index(drop=True)
            anomaly_df = full_df[full_df["_異常"]].drop(columns=["_異常"]).reset_index(drop=True)

            st.subheader("② 科目／級別分組設定")
            st.caption(
                "系統已自動把「同一科目但不同級別」拆分成不同組別（例如 數學組(中一) 與 數學組(中五) 分開，"
                "SS/SS2/SS3 亦已視為三個獨立退修組別：退修1組/退修2組/退修3組）。"
                "您可修改「組別」欄合併多行，或勾選「不填色」跳過該行。"
            )

            edited_normal_df = st.data_editor(
                normal_df,
                num_rows="fixed",
                use_container_width=True,
                disabled=["科目代碼", "級別", "班別示例", "⚠備註"],
                column_config={
                    "組別": st.column_config.TextColumn("組別", help="同一組別會使用同一顏色，可自行合併/改名"),
                    "不填色": st.column_config.CheckboxColumn("不填色", help="勾選代表此行不套用任何顏色"),
                },
                key="normal_mapping_editor",
            )

            if len(anomaly_df) > 0:
                with st.expander(
                    f"⚠ 已自動跳過上色的異常項目共 {len(anomaly_df)} 項（不影響匯出，如需手動處理請展開）",
                    expanded=False,
                ):
                    st.caption("這些項目格式無法明確判斷（例如同一格混合多個級別），系統已預設不上色，可在此手動覆寫。")
                    edited_anomaly_df = st.data_editor(
                        anomaly_df,
                        num_rows="fixed",
                        use_container_width=True,
                        disabled=["科目代碼", "級別", "班別示例", "⚠備註"],
                        column_config={
                            "組別": st.column_config.TextColumn("組別"),
                            "不填色": st.column_config.CheckboxColumn("不填色"),
                        },
                        key="anomaly_mapping_editor",
                    )
            else:
                edited_anomaly_df = anomaly_df

            edited_full_df = pd.concat([edited_normal_df, edited_anomaly_df], ignore_index=True)

            st.subheader("③ 用顏色選擇器為各組別選色")
            active_groups = sorted(set(edited_full_df.loc[~edited_full_df["不填色"], "組別"]))

            if "group_colors" not in st.session_state:
                st.session_state["group_colors"] = {}

            default_colors = assign_default_colors(active_groups)
            for g in active_groups:
                if g not in st.session_state["group_colors"]:
                    st.session_state["group_colors"][g] = default_colors[g]

            st.session_state["group_colors"] = {
                g: c for g, c in st.session_state["group_colors"].items() if g in active_groups
            }

            if not active_groups:
                st.info("目前沒有任何組別需要上色（全部已勾選不填色）。")
            else:
                n_cols = 4
                cols = st.columns(n_cols)
                for i, g in enumerate(active_groups):
                    with cols[i % n_cols]:
                        st.session_state["group_colors"][g] = st.color_picker(
                            g, value=st.session_state["group_colors"][g], key=f"color_{g}"
                        )

            st.subheader("④ 選擇輸出版面格式")
            layout_choice = st.radio(
                "版面格式",
                options=["保留原有格式（不改動版面設置）", "橫向列印格式（A4 Landscape，適合直接列印）"],
                index=0,
                horizontal=True,
                help=(
                    "橫向列印格式會自動設定：列印方向為橫向、縮放至一頁、調整邊界、凍結標題列，"
                    "並將全部字型統一為 Times New Roman、字體大小 12pt，欄寬統一為 16。"
                    "「保留原有格式」則完全不改動版面與字型設定。"
                ),
            )
            layout_mode = "landscape" if layout_choice.startswith("橫向") else "original"

            st.subheader("⑤ 產生彩色 Excel")
            if st.button("🚀 產生並下載彩色時間表", type="primary"):
                try:
                    wb_out = apply_colors_and_export(
                        wb, sheet_names, edited_full_df, st.session_state["group_colors"],
                        layout_mode=layout_mode,
                    )
                    buf = io.BytesIO()
                    wb_out.save(buf)
                    buf.seek(0)
                    st.success(f"已完成！共匯出 {len(sheet_names)} 位老師的彩色時間表 + 顏色對照表"
                               f"（{'橫向列印格式：Times New Roman 12pt，欄寬16' if layout_mode=='landscape' else '原有格式'}）。")
                    st.download_button(
                        "📥 下載彩色時間表 Excel",
                        data=buf,
                        file_name="彩色時間表.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"產生過程發生錯誤：{e}")
else:
    st.info("請先上載一份時間表 Excel 檔案（支援多位老師分頁或單一老師分頁格式）。")
